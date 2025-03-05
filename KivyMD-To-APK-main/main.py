import kivy
from kivy.app import App
from kivy.uix.gridlayout import GridLayout
from kivy.uix.label import Label
from kivy.uix.textinput import TextInput
from kivy.uix.button import Button
from kivy.uix.popup import Popup
from kivy.uix.boxlayout import BoxLayout
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from typing import Optional, Dict, List, Union
import os

kivy.require('2.0.0')  # Replace with your Kivy version if needed

class Pessoa:
    def __init__(self, nome: str, idade: int, sexo: str, cargo: str, abdominal: Optional[int] = None,
                 flexao: Optional[int] = None, corrida: Optional[int] = None):
        self.nome = nome
        self.idade = idade
        self.sexo = sexo
        self.cargo = cargo
        self.abdominal = abdominal
        self.flexao = flexao
        self.corrida = corrida

    def __repr__(self):
        return f"Pessoa(nome='{self.nome}', idade={self.idade}, sexo='{self.sexo}', cargo='{self.cargo}')"

pessoas: List[Pessoa] = []
pessoa_selecionada: Optional[Pessoa] = None
dados_pessoa: Dict[str, Dict[str, Optional[int]]] = {}

class HomePage(GridLayout):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.cols = 1
        self.padding = 20
        self.spacing = 10

        self.welcome_label = Label(text="Bem-vindo ao TAF App!", font_size=40)
        self.add_widget(self.welcome_label)

        self.start_button = Button(text="Iniciar", size_hint=(None, None), size=(200, 50))
        self.start_button.bind(on_press=self.open_selecao_taf)
        self.add_widget(self.start_button)

    def open_selecao_taf(self, instance):
        self.parent.screen_manager.current = 'selecao_taf'

class SelecaoTAFPage(GridLayout):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.cols = 1
        self.padding = 20
        self.spacing = 10

        self.title_label = Label(text="Selecione o Tipo de TAF", font_size=24)
        self.add_widget(self.title_label)

        self.convencional_button = Button(text="TAF Convencional", size_hint=(None, None), size=(300, 70))
        self.convencional_button.bind(on_press=self.open_cadastro)
        self.add_widget(self.convencional_button)

        self.especializado_button = Button(text="TAF Especializado", size_hint=(None, None), size=(300, 70))
        self.especializado_button.bind(on_press=self.open_cadastro)  # Change later
        self.add_widget(self.especializado_button)

    def open_cadastro(self, instance):
        self.parent.screen_manager.current = 'cadastro'

class CadastroPage(GridLayout):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.cols = 1
        self.padding = 20
        self.spacing = 10

        self.nome_label = Label(text="Nome:")
        self.nome_input = TextInput(multiline=False, size_hint_y=None, height=30)
        self.add_widget(self.nome_label)
        self.add_widget(self.nome_input)

        self.idade_label = Label(text="Idade:")
        self.idade_input = TextInput(multiline=False, input_type='number', size_hint_y=None, height=30)
        self.add_widget(self.idade_label)
        self.add_widget(self.idade_input)

        self.sexo_label = Label(text="Sexo:")
        self.sexo_layout = BoxLayout(orientation='horizontal', size_hint_y=None, height=30)
        self.masculino_button = Button(text="Masculino", group="sexo")
        self.feminino_button = Button(text="Feminino", group="sexo")
        self.sexo_layout.add_widget(self.masculino_button)
        self.sexo_layout.add_widget(self.feminino_button)
        self.add_widget(self.sexo_label)
        self.add_widget(self.sexo_layout)

        self.cargo_label = Label(text="Cargo:")
        self.cargo_input = TextInput(multiline=False, size_hint_y=None, height=30)
        self.add_widget(self.cargo_label)
        self.add_widget(self.cargo_input)

        self.cadastrar_button = Button(text="Cadastrar", size_hint_y=None, height=50)
        self.cadastrar_button.bind(on_press=self.cadastrar_pessoa)
        self.add_widget(self.cadastrar_button)

    def cadastrar_pessoa(self, instance):
        nome = self.nome_input.text
        idade = self.idade_input.text
        sexo = None
        for child in self.sexo_layout.children:
            if child.state == 'down':  # Check if the button is selected
                sexo = child.text
                break  # Only one sexo can be selected
        cargo = self.cargo_input.text

        if not all([nome, idade, sexo, cargo]):
            self.show_popup("Por favor, preencha todos os campos.")
            return

        try:
            idade = int(idade)
        except ValueError:
            self.show_popup("Idade deve ser um número inteiro.")
            return

        pessoa = Pessoa(nome, idade, sexo, cargo)
        global pessoas
        pessoas.append(pessoa)

        print(f"Pessoa cadastrada: {pessoa}")
        self.parent.screen_manager.current = 'lista'

        self.nome_input.text = ""
        self.idade_input.text = ""
        for child in self.sexo_layout.children:
            child.state = 'normal'  # Deselect buttons after selection
        self.cargo_input.text = ""
        self.show_popup("Usuário cadastrado com sucesso")

    def show_popup(self, text):
        popup = Popup(title='Aviso',
                      content=Label(text=text),
                      size_hint=(None, None), size=(400, 200))
        popup.open()

class ListaPage(GridLayout):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.cols = 1
        self.padding = 20
        self.spacing = 10

        self.title_label = Label(text="Lista de Pessoas", font_size=24)
        self.add_widget(self.title_label)

        self.add_button = Button(text="Adicionar Candidato", size_hint_y=None, height=50)
        self.add_button.bind(on_press=self.open_cadastro)
        self.add_widget(self.add_button)

        self.search_label = Label(text="Buscar:")
        self.search_input = TextInput(multiline=False, size_hint_y=None, height=30)
        self.search_input.bind(text=self.update_lista_pessoas)
        self.add_widget(self.search_label)
        self.add_widget(self.search_input)

        self.data_button = Button(text="Ver Dados", size_hint_y=None, height=50)
        self.data_button.bind(on_press=self.open_dados_todos)
        self.add_widget(self.data_button)

        self.people_layout = BoxLayout(orientation='vertical', size_hint_y=None)
        self.people_layout.bind(minimum_height=self.people_layout.setter('height'))  # Required for scrolling

        self.add_widget(self.people_layout)
        self.update_lista_pessoas()

    def open_cadastro(self, instance):
        self.parent.screen_manager.current = 'cadastro'

    def open_dados_todos(self, instance):
         self.parent.screen_manager.current = 'dados_todos'

    def update_lista_pessoas(self, instance=None):
         global pessoas

         self.people_layout.clear_widgets()
         search_term = self.search_input.text.lower()

         for pessoa in pessoas:
             if search_term in pessoa.nome.lower() or search_term in pessoa.cargo.lower():
                  pessoa_button = Button(text=pessoa.nome, size_hint_y=None, height=40)
                  pessoa_button.bind(on_press=lambda instance, p=pessoa: self.selecionar_pessoa(p))
                  self.people_layout.add_widget(pessoa_button)

    def selecionar_pessoa(self, pessoa):
         global pessoa_selecionada
         pessoa_selecionada = pessoa
         print(f"Pessoa selecionada: {pessoa_selecionada}")
         self.parent.screen_manager.current = 'dados'
         # Update all required objects with new instance

class DadosPage(GridLayout):
    def __init__(self, **kwargs):
         super().__init__(**kwargs)
         self.cols = 1
         self.padding = 20
         self.spacing = 10

         self.abdominal_label = Label(text="Abdominal:")
         self.abdominal_input = TextInput(multiline=False, input_type='number', size_hint_y=None, height=30)
         self.add_widget(self.abdominal_label)
         self.add_widget(self.abdominal_input)

         self.flexao_label = Label(text="Flexão:")
         self.flexao_input = TextInput(multiline=False, input_type='number', size_hint_y=None, height=30)
         self.add_widget(self.flexao_label)
         self.add_widget(self.flexao_input)

         self.corrida_label = Label(text="Corrida:")
         self.corrida_input = TextInput(multiline=False, input_type='number', size_hint_y=None, height=30)
         self.add_widget(self.corrida_label)
         self.add_widget(self.corrida_input)

         self.salvar_button = Button(text="Salvar", size_hint_y=None, height=50)
         self.salvar_button.bind(on_press=self.salvar_dados)
         self.add_widget(self.salvar_button)

    def salvar_dados(self, instance):
         global pessoa_selecionada, dados_pessoa
         if pessoa_selecionada:
              try:
                   abdominal = int(self.abdominal_input.text) if self.abdominal_input.text else None
                   flexao = int(self.flexao_input.text) if self.flexao_input.text else None
                   corrida = int(self.corrida_input.text) if self.corrida_input.text else None
              except ValueError:
                   self.show_popup("Os campos Abdominal, Flexão e Corrida devem ser números inteiros.")
                   return

              dados_pessoa[pessoa_selecionada.nome] = {
                   "abdominal": abdominal,
                   "flexao": flexao,
                   "corrida": corrida,
              }

              pessoa_selecionada.abdominal = abdominal
              pessoa_selecionada.flexao = flexao
              pessoa_selecionada.corrida = corrida

              print(f"Dados salvos para {pessoa_selecionada.nome}: {dados_pessoa[pessoa_selecionada.nome]}")

              self.parent.screen_manager.current = 'lista'

              self.abdominal_input.text = ""
              self.flexao_input.text = ""
              self.corrida_input.text = ""
         else:
              self.show_popup("Nenhuma pessoa selecionada para salvar os dados.")

    def show_popup(self, text):
         popup = Popup(title='Aviso',
                      content=Label(text=text),
                      size_hint=(None, None), size=(400, 200))
         popup.open()

class DadosTodosPage(GridLayout):

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.cols = 1
        self.padding = 20
        self.spacing = 10

        self.title_label = Label(text="Dados dos Candidatos", font_size=24)
        self.add_widget(self.title_label)

        self.data_layout = GridLayout(cols=7, size_hint_y=None)  # Layout for the table data
        self.data_layout.bind(minimum_height=self.data_layout.setter('height'))  # For scrolling

        # Create headers
        headers = ["Nome", "Idade", "Sexo", "Cargo", "Abdominal", "Flexão", "Corrida"]
        for header in headers:
            self.data_layout.add_widget(Label(text=header, bold=True))

        self.update_table()

        self.add_widget(self.data_layout)

        self.export_button = Button(text="Exportar para Excel", size_hint_y=None, height=50)
        self.export_button.bind(on_press=self.export_to_excel)
        self.add_widget(self.export_button)

    def update_table(self):
        global pessoas
        self.data_layout.clear_widgets()
        # Create headers
        headers = ["Nome", "Idade", "Sexo", "Cargo", "Abdominal", "Flexão", "Corrida"]
        for header in headers:
            self.data_layout.add_widget(Label(text=header, bold=True))
        for pessoa in pessoas:
            self.data_layout.add_widget(Label(text=pessoa.nome))
            self.data_layout.add_widget(Label(text=str(pessoa.idade)))
            self.data_layout.add_widget(Label(text=pessoa.sexo))
            self.data_layout.add_widget(Label(text=pessoa.cargo))
            self.data_layout.add_widget(Label(text=str(pessoa.abdominal) or ""))
            self.data_layout.add_widget(Label(text=str(pessoa.flexao) or ""))
            self.data_layout.add_widget(Label(text=str(pessoa.corrida) or ""))

    def export_to_excel(self, instance):
        global pessoas
        wb = Workbook()
        ws = wb.active
        colunas = ["Nome", "Idade", "Sexo", "Cargo", "Abdominal", "Flexão", "Corrida"]

        # Escreve os cabeçalhos
        for col_num, column_title in enumerate(colunas, 1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}1"] = column_title

        # Escreve os dados
        for row_num, pessoa in enumerate(pessoas, 2):
            ws[f"A{row_num}"] = pessoa.nome
            ws[f"B{row_num}"] = pessoa.idade
            ws[f"C{row_num}"] = pessoa.sexo
            ws[f"D{row_num}"] = pessoa.cargo
            ws[f"E{row_num}"] = pessoa.abdominal
            ws[f"F{row_num}"] = pessoa.flexao
            ws[f"G{row_num}"] = pessoa.corrida

        # Ajustar a largura das colunas
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2

        # Salvar o arquivo
        file_path = "dados_taf.xlsx"
        wb.save(file_path)
        self.show_popup("Tabela salva com sucesso")

    def show_popup(self, text):
        popup = Popup(title='Aviso',
                      content=Label(text=text),
                      size_hint=(None, None), size=(400, 200))
        popup.open()

from kivy.uix.screenmanager import ScreenManager, Screen

class HomePageScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.name = 'home'
        home_page = HomePage()
        home_page.parent = self  # Set the parent to the screen
        self.add_widget(home_page)

class SelecaoTAFScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.name = 'selecao_taf'
        selecao_taf_page = SelecaoTAFPage()
        selecao_taf_page.parent = self  # Set the parent to the screen
        self.add_widget(selecao_taf_page)

class CadastroScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.name = 'cadastro'
        cadastro_page = CadastroPage()
        cadastro_page.parent = self  # Set the parent to the screen
        self.add_widget(cadastro_page)

class ListaScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.name = 'lista'
        lista_page = ListaPage()
        lista_page.parent = self
        self.add_widget(lista_page)

class DadosScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.name = 'dados'
        dados_page = DadosPage()
        dados_page.parent = self
        self.add_widget(dados_page)

class DadosTodosScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.name = 'dados_todos'
        dados_todos_page = DadosTodosPage()
        dados_todos_page.parent = self
        self.add_widget(dados_todos_page)

class TAFApp(App):
    def build(self):
        self.screen_manager = ScreenManager()

        self.home_page_screen = HomePageScreen()
        self.selecao_taf_screen = SelecaoTAFScreen()
        self.cadastro_screen = CadastroScreen()
        self.lista_screen = ListaScreen()
        self.dados_screen = DadosScreen()
        self.dados_todos_screen = DadosTodosScreen()

        self.screen_manager.add_widget(self.home_page_screen)
        self.screen_manager.add_widget(self.selecao_taf_screen)
        self.screen_manager.add_widget(self.cadastro_screen)
        self.screen_manager.add_widget(self.lista_screen)
        self.screen_manager.add_widget(self.dados_screen)
        self.screen_manager.add_widget(self.dados_todos_screen)

        return self.screen_manager

if __name__ == '__main__':
    TAFApp().run()
